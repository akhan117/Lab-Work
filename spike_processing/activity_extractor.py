import h5py
from tkinter.filedialog import *
import ntpath
import pickle
from openpyxl import Workbook
from os.path import exists

# Written by Ayaan Khan
# This program uses the files created by event_spike_merger, and depending on user input creates an Excel file
# with the spike data from Spyking-Circus

# Python alphabet lists that go up to 'ZZ' eventually, for convenience.
letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
           'V', 'W', 'X', 'Y', 'Z']

lettersExpanded = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                   'U', 'V', 'W', 'X', 'Y', 'Z']

# Creating our 'ZZ' list
for i in letters:
    for j in letters:
        lettersExpanded.append(i + j)

# This variable exists to check if we are in the first loop
omega_check = 0
if __name__ == "__main__":
    # The user is able to consecutively choose files, until they cancel the process.
    print("Select the files you'd like to extract Data from")
    wb = Workbook()

    OdorList = set()
    fileList = []
    while True:
        val = askopenfilename()
        fileList.append(val)

        if val == "":
            break
    fileList = fileList[:-1]
    megaCount = 1

    # By knowing Spyking-Circus naming conventions, we deduce the name of the file whose data we need
    for file in fileList:
        file2 = ntpath.basename(file)
        file2 = file2[:-19]
        file2 = file2.split(",")[-1]
        file2 = file2.split()[1:]
        fileNew = ""

        for i in file2:
            fileNew = fileNew + i + " "
        # We obtain events, splits, sampling rates and timing data
        folder = "Event and Sampling Rates Data/" + fileNew
        splits = folder + "splits.pk"
        with open(splits, 'rb') as fi:
            splits_list = pickle.load(fi)

        rate_file = "Default Values/Default Sampling Rate.pk"
        with open(rate_file, 'rb') as fi:
            sampling = pickle.load(fi)

        events = folder + "events.pk"
        with open(events, 'rb') as fi:
            eventList = pickle.load(fi)

        times = folder + "times.pk"
        with open(times, 'rb') as fi:
            timeList = pickle.load(fi)

        # We recreate our events + time list for convenience
        event_time = []
        for i in range(0, len(eventList)):
            event_time.append((eventList[i], timeList[i]))

        # We figure out which data is pre infusion and which is post infusion.
        splits_list = splits_list[:-1]
        splits_list.insert(0, 0.0)
        splitn = splits_list[1] / sampling

        # We use the file we chose earlier and extract the spikes timings, and the odor onsets and offsets
        neurons = []
        names, onsets, offsets = '', '', ''
        with h5py.File(file, 'r+') as f:
            names = f["Odor_Names"][:]
            onsets = f["Odor_Onsets"][:]
            offsets = f["Odor_Offsets"][:]
            neuron_test = f["Neurons"]
            for i in neuron_test:
                neurons.append(neuron_test[i][:])

        timingList = []
        for i in range(0, len(names)):
            timingList.append((names[i], onsets[i]))

        #  The program splits depending on what the user needs to do
        if megaCount == 1:
            print("Type in the number of the option you'd like to run -")
            print("1) Spontaneous Activity before any Odors")
            print("2) Odor responses")
            option = input()

        # #############################################################################################################
        if option == '1':
            # Data is collected for "Spontaneous Activity"
            if megaCount == 1:
                print("What is the start point of the data?")
                start = float(input())
                postStart = start + splitn
                print("How many seconds of data do you want?")
                length = float(input())

            # Create a dictionary for every neuron, pre- and post- infusion to store our data
            neuro_pre = {}
            neuro_post = {}
            count = 1
            for n in neurons:
                name = "neuron_" + str(count)
                neuro_pre[name] = []
                neuro_post[name] = []
                count += 1

            count = 1
            for n in neurons:
                name = "neuron_" + str(count)
                ls = []
                ls2 = []
                for n1 in n:
                    if start < n1 < start + length:
                        ls.append(n1)

                    elif postStart < n1 < postStart + length:
                        ls2.append(n1)
                neuro_pre[name].append(len(ls))
                neuro_post[name].append(len(ls2))
                count += 1
            print("Pre Drugs:" + str(neuro_pre))
            print("Post Drugs:" + str(neuro_post))
            print()

            # If save to location exists, save Excel file there, otherwise ask the user
            if exists("Default Values/Default Save to Location E.pk"):
                print()
                print(
                    "Using the Default save to location from Default Save to Location E.pk! Delete this file if you do"
                    + " not want to use the default save to location anymore!")

                with open("Default Values/Default Save to Location E.pk", 'rb') as fi:
                    excelName = pickle.load(fi)

                print("Saving to " + excelName)

            else:
                print(
                    "#################################################################################################")
                print("Where do you want to save the files to")
                save_to = askdirectory()
                print("What do you want to name the file?")
                print(
                    "#################################################################################################")

                excelName2 = input()
                excelName = (save_to + '\\' + excelName2 + ".xlsx")

                print("Do you want save this as the default location to save to (y/n)")
                default = input()

                if default == 'y':
                    with open("Default Values/Default Save to Location E.pk", 'wb') as fi:
                        pickle.dump(excelName, fi)

                elif default == 'n':
                    print()

                else:
                    print("Defaulting to n.")

            # Save our data to the Excel file, how depends on whether this is the first loop
            miniCount = str(megaCount)
            ws = wb.active
            if omega_check == 0:
                ws['A' + miniCount] = "File Name"
                ws['D' + miniCount] = "Neuron ID"
                ws["B" + miniCount] = "Start Time"
                ws["C" + miniCount] = "Duration"
                ws["E" + miniCount] = "Pre Infusion Spikes"
                ws["F" + miniCount] = "Post Infusion Spikes"
                omega_check = 1

            counter = megaCount + 1
            for i in neuro_pre:
                ws['A' + str(counter)] = ntpath.basename(file)
                ws["B" + str(counter)] = start
                ws["C" + str(counter)] = length
                ws['D' + str(counter)] = i
                ws['E' + str(counter)] = neuro_pre[i][0]
                ws['F' + str(counter)] = neuro_post[i][0]
                counter += 1

            wb.save(excelName)
            megaCount += len(neurons)



        ################################################################################################################
        elif option == '2':
            # Collect data we need to create the bins for Odor Responses.
            if megaCount == 1:
                print("Enter the number of seconds before odor infusions you'd like to analyze.")
                timeB = float(input())
                print("Enter the number of seconds after odor infusions you'd like to analyze.")
                timeA = float(input())
                print("Enter the size of each bin.")
                binSize = float(input())

                binNoB = int(timeB / binSize)
                binNoA = int(timeA / binSize)
                print("The Number of bins before each odor infusions is " + str(binNoB))
                print("The Number of bins after each odors infusions is " + str(binNoA))

                # If save to location exists, save Excel file there, otherwise ask the user
                if exists("Default Values/Default Save to Location E2.pk"):
                    print()
                    print(
                        "Using the Default save to location from Default Save to Location E2.pk! Delete this file if "
                        + "you do not want to use the default save to location anymore!")

                    with open("Default Values/Default Save to Location E2.pk", 'rb') as fi:
                        excelName = pickle.load(fi)

                    print("Saving to " + excelName)

                else:
                    print(
                        "#############################################################################################")
                    print("Where do you want to save the files to")
                    save_to = askdirectory()
                    print("What do you want to name the file?")
                    print(
                        "#############################################################################################")

                    excelName2 = input()
                    excelName = (save_to + '\\' + excelName2 + ".xlsx")

                    print("Do you want save this as the default location to save to (y/n)")
                    default = input()

                    if default == 'y':
                        with open("Default Values/Default Save to Location E2.pk", 'wb') as fi:
                            pickle.dump(excelName, fi)

                    elif default == 'n':
                        print()

                    else:
                        print("Defaulting to n.")

            odorData = {}
            for n in names:
                OdorList.add(n)
                odorData[n] = []
            # Create the number of bins we need
            binDivB = []
            binDivA = []
            for i in range(1, binNoB + 1):
                binDivB.append(binSize * i)

            for i in range(1, binNoA + 1):
                binDivA.append(binSize * i)

            ws = wb.active

            # Save the exact timings of the bins
            binTimingsPre = []
            binTimingsPost = []
            for i in event_time:
                time = i[1]
                temp = []
                temp2 = []
                for j in binDivB:
                    temp.append(time - j)

                temp.append(time)
                temp.sort()
                temp2.append(time)
                for k in binDivA:
                    temp2.append(time + k)

                binTimingsPre.append((i[0], temp))
                binTimingsPost.append((i[0], temp2))

            # pr = Pre, p = Post, I = Infusion, O = Odor. e.g - prIpO would be pre Infusion, post Odor dictionary
            prIprO = {}
            pIprO = {}
            prIpO = {}
            pIpO = {}

            count = 1

            # for each neuron, create an entry in our dictionaries, and fill the entry in with the number of spikes
            # in each bin
            for n1 in neurons:
                neuron_ID = "neuron_" + str(count)
                prIprO[neuron_ID] = []
                pIprO[neuron_ID] = []
                prIpO[neuron_ID] = []
                pIpO[neuron_ID] = []
                full_list_prIprO = []
                full_list_pIprO = []
                full_list_prIpO = []
                full_list_pIpO = []

                count += 1
                # For the pre infusion bins
                for [odor, timing] in binTimingsPre:
                    temp1 = [odor]
                    temp2 = [odor]

                    for t in range(0, len(timing) - 1):

                        count1 = 0
                        count2 = 0
                        updated_split = timing[len(timing) - 1]
                        for n in n1:

                            if timing[t] < n < timing[t + 1]:
                                if n < splitn + timeB:
                                    count1 += 1
                                else:
                                    count2 += 1

                        if timing[t] < splitn + timeB:
                            temp1.append(count1)
                        else:
                            temp2.append(count2)

                    if len(temp1) == binNoB + 1:
                        full_list_prIprO.append(temp1)
                    if len(temp2) == binNoB + 1:
                        full_list_pIprO.append(temp2)
                prIprO[neuron_ID] = full_list_prIprO
                pIprO[neuron_ID] = full_list_pIprO

                # for the post infusion bins
                for [odor, timing] in binTimingsPost:
                    temp1 = [odor]
                    temp2 = [odor]

                    for t in range(0, len(timing) - 1):

                        count1 = 0
                        count2 = 0

                        for n in n1:
                            if timing[t] < n < timing[t + 1]:

                                if n < splitn + timeA:
                                    count1 += 1
                                else:
                                    count2 += 1

                        if timing[t] < splitn + timeA:
                            temp1.append(count1)
                        else:
                            temp2.append(count2)

                    if len(temp1) == binNoA + 1:
                        full_list_prIpO.append(temp1)
                    if len(temp2) == binNoA + 1:
                        full_list_pIpO.append(temp2)

                prIpO[neuron_ID] = full_list_prIpO
                pIpO[neuron_ID] = full_list_pIpO

            miniCount = str(megaCount)
            # If first run, fill out the column titles
            if omega_check == 0:
                ws['A' + miniCount] = "File Name"
                ws['B' + miniCount] = "Pre Odor Bins"
                ws['C' + miniCount] = "Post Odor Bins"
                ws['D' + miniCount] = "Bin Size"
                ws['E' + miniCount] = "Odor Identity"
                ws['F' + miniCount] = "Neuron_ID"

                clet = lettersExpanded.index('G')
                count = 1
                for i in range(0, binNoB):
                    ws[lettersExpanded[clet] + miniCount] = "PriPrO Bin " + str(count)
                    clet += 1
                    count += 1

                count = 1
                for i in range(0, binNoA):
                    ws[lettersExpanded[clet] + miniCount] = "PriPO Bin " + str(count)
                    clet += 1
                    count += 1

                ws[lettersExpanded[clet] + miniCount] = "Odor Identity"
                clet += 1
                ws[lettersExpanded[clet] + miniCount] = "Neuron_ID"
                clet += 1

                count = 1
                for i in range(0, binNoB):
                    ws[lettersExpanded[clet] + miniCount] = "PiPrO Bin " + str(count)
                    clet += 1
                    count += 1

                count = 1
                for i in range(0, binNoA):
                    ws[lettersExpanded[clet] + miniCount] = "PiPO Bin " + str(count)
                    clet += 1
                    count += 1

                omega_check = 1

            # We work through our four dictionaries and print it into the Excel files
            something = megaCount + 1
            something1 = something
            something2 = something
            something3 = something
            keys = prIprO.keys()
            keys = sorted(keys)
            buffer = len(keys) + 1

            for i in keys:
                count = 0
                for j in prIprO[i]:
                    ws['E' + str(something)] = j[0]
                    ws['F' + str(something)] = i
                    clet = lettersExpanded.index('G')
                    for k in j[1:]:
                        ws[lettersExpanded[clet] + str(something)] = k
                        clet += 1

                    ws['A' + str(something)] = ntpath.basename(file)
                    ws['B' + str(something1)] = binNoB
                    ws['C' + str(something1)] = binNoA
                    ws['D' + str(something1)] = binSize
                    something += 1

                ct = clet
                for j in prIpO[i]:
                    ws['E' + str(something1)] = j[0]
                    ws['F' + str(something1)] = i
                    clet = ct
                    for k in j[1:]:
                        ws[lettersExpanded[clet] + str(something1)] = k
                        clet += 1

                    ws['A' + str(something1)] = ntpath.basename(file)
                    ws['B' + str(something1)] = binNoB
                    ws['C' + str(something1)] = binNoA
                    ws['D' + str(something1)] = binSize
                    something1 += 1

                ct = clet
                for j in pIprO[i]:
                    clet = ct
                    ws[lettersExpanded[clet] + str(something2)] = j[0]
                    clet += 1

                    ws[lettersExpanded[clet] + str(something2)] = i
                    clet += 1

                    for k in j[1:]:
                        ws[lettersExpanded[clet] + str(something2)] = k
                        clet += 1

                    ws['A' + str(something2)] = ntpath.basename(file)
                    ws['B' + str(something2)] = binNoB
                    ws['C' + str(something2)] = binNoA
                    ws['D' + str(something2)] = binSize
                    something2 += 1

                ct = clet
                for j in pIpO[i]:
                    clet = ct
                    for k in j[1:]:
                        ws[lettersExpanded[clet] + str(something3)] = k
                        clet += 1

                    something3 += 1

            something = max(something, something1, something2, something3)
            megaCount = something - 1
            wb.save(excelName)

    print("Finished")
